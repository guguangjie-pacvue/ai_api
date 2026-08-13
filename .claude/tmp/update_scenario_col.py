"""
Scan all latest cases.json, extract per-path scenario labels + percentages,
then write them into the '场景覆盖' column of each swagger sheet in Excel.
Format per cell:
  场景1: <desc> — xx%
  场景2: <desc> — xx%
  ...
"""
import glob, json, re, os
from collections import defaultdict
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

EXCEL_PATH = 'C:/AI engineering/single-api/ai_api/single-api/swagger_modules.xlsx'
BASE = 'C:/AI engineering/single-api/ai_api/single-api'

# ── 1. Find latest task per (swagger, module) ────────────────────────────────
task_dirs = {}
for f in glob.glob(BASE + '/mainapi/**/**/task-*/cases.json', recursive=True):
    fp = f.replace('\\', '/')
    rel = fp.split('/mainapi/')[1]
    parts = rel.split('/')
    if len(parts) < 4:
        continue
    swagger, module, task = parts[0], parts[1], parts[2]
    key = (swagger, module)
    if key not in task_dirs or task > task_dirs[key][0]:
        task_dirs[key] = (task, f)

# ── 2. Build path → scenario lines ──────────────────────────────────────────
# Key: (swagger_sheet_name, full_path_with_api_prefix)  e.g. ('Amazon.Advertising.Api', '/api/Target/GetXxx')
# Value: list of "场景N: desc — xx%"
scenario_map = defaultdict(list)

def clean_scenario_label(raw):
    """Convert '场景1_Campaign+日期过滤_xxx' → '场景1: Campaign+日期过滤 xxx'"""
    # split off 场景N prefix
    m = re.match(r'^(场景\d+)[_\-](.+)$', raw)
    if m:
        no, rest = m.group(1), m.group(2)
        rest = rest.replace('_', ' ')
        return no + ': ' + rest
    return raw.replace('_', ' ')

for (swagger, module), (task, fpath) in sorted(task_dirs.items()):
    try:
        with open(fpath, encoding='utf-8-sig') as fh:
            cases = json.load(fh)
    except Exception as e:
        print(f'  SKIP {fpath}: {e}')
        continue

    # group by path
    path_cases = defaultdict(list)
    for c in cases:
        steps = c.get('steps', [])
        if not steps:
            continue
        raw_path = steps[0].get('path', '')
        full_path = '/api/' + raw_path if not raw_path.startswith('/') else raw_path
        name = c.get('name', '')
        desc = c.get('description', '')

        # extract scenario label from name (after ' - ')
        if ' - ' in name:
            label_raw = name.split(' - ', 1)[1]
        else:
            label_raw = name

        label = clean_scenario_label(label_raw)

        # extract percentage from description
        pct_m = re.search(r'占比约([\d.]+%)', desc)
        pct = pct_m.group(1) if pct_m else '?%'

        path_cases[full_path].append((label, pct))

    for full_path, items in path_cases.items():
        lines = [f'{label} — {pct}' for label, pct in items]
        scenario_map[(swagger, full_path)] = lines

print(f'Built scenario_map: {len(scenario_map)} entries')
for k, v in sorted(scenario_map.items()):
    print(f'  {k[0]} | {k[1]}: {len(v)} scenarios')

# ── 3. Update Excel ──────────────────────────────────────────────────────────
thin  = Side(style='thin', color='BFBFBF')
bdr   = Border(left=thin, right=thin, top=thin, bottom=thin)
header_font = Font(bold=True, color='FFFFFF')
header_fill = PatternFill('solid', fgColor='2E5F8A')
COL_NAME = '场景覆盖'
COL_WIDTH = 72

wb = openpyxl.load_workbook(EXCEL_PATH)

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    # Only process swagger-level sheets (skip 模块汇总)
    if sheet_name not in ('Amazon.Advertising.Api', 'PacvueMainApi'):
        continue

    # Find path column index
    path_col = None
    for c in range(1, ws.max_column + 1):
        v = str(ws.cell(1, c).value or '')
        if '路径' in v or 'path' in v.lower():
            path_col = c
            break
    if path_col is None:
        print(f'[{sheet_name}] path column not found, skip')
        continue

    # Find or create 场景覆盖 column
    scene_col = None
    for c in range(1, ws.max_column + 1):
        if ws.cell(1, c).value == COL_NAME:
            scene_col = c
            break
    if scene_col is None:
        scene_col = ws.max_column + 1
        hcell = ws.cell(1, scene_col, COL_NAME)
        hcell.font      = header_font
        hcell.fill      = header_fill
        hcell.alignment = Alignment(horizontal='center', vertical='center')
        hcell.border    = bdr
        ws.column_dimensions[get_column_letter(scene_col)].width = COL_WIDTH
        print(f'[{sheet_name}] Created column "{COL_NAME}" at col {scene_col}')

    updated = 0
    for r in range(2, ws.max_row + 1):
        path = ws.cell(r, path_col).value
        if not path:
            continue
        key = (sheet_name, path)
        if key in scenario_map:
            lines = scenario_map[key]
            text  = '\n'.join(lines)
            cell  = ws.cell(r, scene_col, text)
            cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
            cell.border    = bdr
            ws.row_dimensions[r].height = max(
                ws.row_dimensions[r].height or 15,
                len(lines) * 16
            )
            updated += 1

    print(f'[{sheet_name}] Updated {updated} rows')

wb.save(EXCEL_PATH)
print('Done. Excel saved.')
