import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

EXCEL_PATH = 'C:/AI engineering/single-api/ai_api/single-api/swagger_modules.xlsx'

# Scenario data: path -> list of "场景N: desc (xx%)"
SCENARIO_MAP = {
    '/api/SupplementData/GetSupplementTableData': [
        '场景1: Campaign+日期过滤 → Campaign汇总(含AdGroupCount) — 52.6%',
        '场景2: Campaign+日期过滤 → CostControl输出(无AdGroupCount) — 18.8%',
        '场景3: Campaign+AdGroup+ASIN → ASIN评分/品牌数据 — 13.6%',
        '场景4: 仅Campaign过滤 → 受众数据(AudienceId/AudienceResolved) — 7.4%',
        '场景5: Campaign+AdGroup+日期 → 广告组运营数据 — 3.0%',
        '场景6: Campaign+日期+IsFead=0 → 花费/销售占比数据 — 2.4%',
    ]
}

thin  = Side(style='thin', color='BFBFBF')
bdr   = Border(left=thin, right=thin, top=thin, bottom=thin)

header_font = Font(bold=True, color='FFFFFF')
header_fill = PatternFill('solid', fgColor='2E5F8A')

wb = openpyxl.load_workbook(EXCEL_PATH)

for sheet_name in ['Amazon.Advertising.Api', 'PacvueMainApi']:
    ws = wb[sheet_name]

    # Find or create 场景覆盖 column
    col_idx = None
    for c in range(1, ws.max_column + 1):
        if ws.cell(1, c).value == '场景覆盖':
            col_idx = c
            break
    if col_idx is None:
        col_idx = ws.max_column + 1
        hcell = ws.cell(1, col_idx, '场景覆盖')
        hcell.font      = header_font
        hcell.fill      = header_fill
        hcell.alignment = Alignment(horizontal='center', vertical='center')
        hcell.border    = bdr
        from openpyxl.utils import get_column_letter
        ws.column_dimensions[get_column_letter(col_idx)].width = 72

    # Fill scenario data
    path_col = None
    for c in range(1, ws.max_column + 1):
        v = ws.cell(1, c).value or ''
        if '路径' in str(v) or 'path' in str(v).lower():
            path_col = c
            break
    if path_col is None:
        print(f'[{sheet_name}] path column not found, skip')
        continue

    updated = 0
    for r in range(2, ws.max_row + 1):
        path = ws.cell(r, path_col).value
        if path in SCENARIO_MAP:
            text = '\n'.join(SCENARIO_MAP[path])
            cell = ws.cell(r, col_idx, text)
            cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
            cell.border    = bdr
            ws.row_dimensions[r].height = max(ws.row_dimensions[r].height or 15,
                                              len(SCENARIO_MAP[path]) * 16)
            updated += 1
            print(f'[{sheet_name}] Row {r} updated: {path}')

    if updated == 0:
        print(f'[{sheet_name}] No matching paths found')

wb.save(EXCEL_PATH)
print('Done.')
