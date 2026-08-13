
import json, os
from pathlib import Path

base = Path('C:/AI engineering/single-api/ai_api/single-api')

# Load endpoints files for api_total per module
endpoints = {}
for ep_file in base.glob('endpoints-*.json'):
    title = ep_file.stem.replace('endpoints-', '')
    with open(ep_file, encoding='utf-8') as f:
        data = json.load(f)
    for item in data:
        tag = item['tag']
        key = (title, tag)
        endpoints[key] = endpoints.get(key, 0) + 1

# Find latest report.json per (swagger, module)
reports = {}
for rp in sorted(base.rglob('report.json')):
    parts = rp.parts
    try:
        idx = [i for i, p in enumerate(parts) if p == 'mainapi'][0]
        swagger = parts[idx+1]
        module  = parts[idx+2]
        task    = parts[idx+3]
        key = (swagger, module)
        if key not in reports or task > reports[key][1]:
            reports[key] = (rp, task)
    except:
        pass

results = {}
for (swagger, module), (rp, task) in sorted(reports.items()):
    with open(rp, encoding='utf-8') as f:
        r = json.load(f)

    cases_path = rp.parent / 'cases.json'
    api_covered = 0
    if cases_path.exists():
        with open(cases_path, encoding='utf-8') as f:
            cases = json.load(f)
        paths = set()
        for c in cases:
            for s in c.get('steps', []):
                p = s.get('path', '')
                if p:
                    paths.add(p.split('?')[0])  # 去掉 query 参数再去重
        api_covered = len(paths)

    # 规范化 case 路径，与 swagger 路径对齐（统一去前缀后仅取末段接口名比较）
    def normalize(p):
        p = p.split('?')[0].strip('/')
        return p.split('/')[-1].lower()

    case_endpoints = set(normalize(p) for p in paths)

    api_total = endpoints.get((swagger, module), 0)
    # 用 swagger 路径做分母，计算有多少被 case 覆盖
    swagger_eps = [e['path'] for e in __import__('json').load(open(
        'C:/AI engineering/single-api/ai_api/single-api/endpoints-' + swagger + '.json', encoding='utf-8'))
        if e['tag'] == module]
    covered = sum(1 for ep in swagger_eps if normalize(ep) in case_endpoints)
    api_covered = covered

    pass_rate = (str(round(r['passed'] / r['total'] * 100)) + '%') if r['total'] else 'N/A'
    if api_total:
        coverage = str(round(api_covered / api_total * 100)) + '% (' + str(api_covered) + '/' + str(api_total) + ')'
    else:
        coverage = '?(' + str(api_covered) + '/??)'

    results[(swagger, module)] = {
        'case': r['total'],
        'passed': r['passed'],
        'pass_rate': pass_rate,
        'coverage': coverage,
    }
    print(swagger, '|', module, '| case=' + str(r['total']), 'passed=' + str(r['passed']), '|', pass_rate, '|', coverage)

# Write to Excel
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

EXCEL_PATH     = 'C:/AI engineering/single-api/ai_api/single-api/swagger_modules.xlsx'
EXCEL_PATH_TMP = 'C:/AI engineering/single-api/ai_api/single-api/swagger_modules_tmp.xlsx'
wb = openpyxl.load_workbook(EXCEL_PATH)
ws = wb['模块汇总']

headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
new_cols = {'Case数': None, '通过率': None, '接口覆盖率': None}
for col_name in new_cols:
    if col_name in headers:
        new_cols[col_name] = headers.index(col_name) + 1
    else:
        col_idx = ws.max_column + 1
        cell = ws.cell(1, col_idx, col_name)
        cell.font = Font(bold=True)
        cell.fill = PatternFill('solid', fgColor='D9E1F2')
        cell.alignment = Alignment(horizontal='center')
        new_cols[col_name] = col_idx

updated = []
for row_idx in range(2, ws.max_row + 1):
    swagger = ws.cell(row_idx, 1).value
    module  = ws.cell(row_idx, 2).value
    key = (swagger, module)
    if key in results:
        d = results[key]
        ws.cell(row_idx, new_cols['Case数'],    d['case']).alignment      = Alignment(horizontal='center')
        ws.cell(row_idx, new_cols['通过率'],    d['pass_rate']).alignment = Alignment(horizontal='center')
        ws.cell(row_idx, new_cols['接口覆盖率'], d['coverage']).alignment  = Alignment(horizontal='center')
        updated.append(module)

import shutil, os
wb.save(EXCEL_PATH_TMP)
# 替换原文件（若原文件被锁则报错）
try:
    shutil.move(EXCEL_PATH_TMP, EXCEL_PATH)
    print('\nUpdated modules:', updated)
    print('Saved to:', EXCEL_PATH)
except PermissionError:
    print('\nUpdated modules:', updated)
    print('Original file locked. Saved to tmp:', EXCEL_PATH_TMP)
    print('Please close swagger_modules.xlsx and rename the tmp file manually.')
