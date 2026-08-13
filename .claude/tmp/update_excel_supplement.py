import openpyxl, json
from openpyxl.styles import Font, PatternFill, Alignment

EXCEL_PATH = 'C:/AI engineering/single-api/ai_api/single-api/swagger_modules.xlsx'
REPORT_PATH = 'C:/AI engineering/single-api/ai_api/single-api/mainapi/Amazon.Advertising.Api/SupplementData/task-2026-08-13-15-32-43/report.json'
ENDPOINTS_PATH = 'C:/AI engineering/single-api/ai_api/single-api/endpoints-Amazon.Advertising.Api.json'
CASES_PATH = 'C:/AI engineering/single-api/ai_api/single-api/mainapi/Amazon.Advertising.Api/SupplementData/task-2026-08-13-15-32-43/cases.json'

with open(REPORT_PATH, encoding='utf-8') as f:
    r = json.load(f)

with open(ENDPOINTS_PATH, encoding='utf-8') as f:
    endpoints = json.load(f)

api_total = len([e for e in endpoints if e['tag'] == 'SupplementData'])

case_paths = set()
with open(CASES_PATH, encoding='utf-8') as f:
    cases = json.load(f)
for c in cases:
    for s in c.get('steps', []):
        case_paths.add(s.get('path', ''))
api_covered = len(case_paths)

results = {('Amazon.Advertising.Api', 'SupplementData'): {
    'case': r['total'], 'passed': r['passed'],
    'api_covered': api_covered, 'api_total': api_total
}}

wb = openpyxl.load_workbook(EXCEL_PATH)
ws = wb['模块汇总']

headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
new_cols = {'Case数': None, '通过率': None, '接口覆盖率': None}
for col_name in new_cols:
    if col_name in headers:
        new_cols[col_name] = headers.index(col_name) + 1
    else:
        col_idx = ws.max_column + 1
        cell = ws.cell(1, col_idx, col_name)
        cell.font = Font(bold=True)
        cell.fill = PatternFill('solid', fgColor='D9E1F2')
        cell.alignment = Alignment(horizontal='center')
        new_cols[col_name] = col_idx

for row_idx in range(2, ws.max_row + 1):
    swagger = ws.cell(row_idx, 1).value
    module  = ws.cell(row_idx, 2).value
    key = (swagger, module)
    if key in results:
        d = results[key]
        pass_rate = str(round(d['passed'] / d['case'] * 100)) + '%' if d['case'] else 'N/A'
        coverage  = str(round(d['api_covered'] / d['api_total'] * 100)) + '% (' + str(d['api_covered']) + '/' + str(d['api_total']) + ')'
        ws.cell(row_idx, new_cols['Case数'],    d['case']).alignment = Alignment(horizontal='center')
        ws.cell(row_idx, new_cols['通过率'],    pass_rate).alignment = Alignment(horizontal='center')
        ws.cell(row_idx, new_cols['接口覆盖率'], coverage).alignment = Alignment(horizontal='center')
        print('Updated row', row_idx, ':', swagger, '/', module, '-> case=', d['case'], 'pass=', pass_rate, 'coverage=', coverage)

wb.save(EXCEL_PATH)
print('Excel updated.')
