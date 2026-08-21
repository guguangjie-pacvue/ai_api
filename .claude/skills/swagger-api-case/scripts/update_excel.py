"""
update_excel.py  —  Phase 5 Excel 写入（模块汇总 + 场景覆盖）

Usage:
  python scripts/update_excel.py \\
    --cases   single-api/mainapi/us/Amazon.Advertising.Api/SupplementData/task-xxx/cases.json \\
    --report  single-api/mainapi/us/Amazon.Advertising.Api/SupplementData/task-xxx/report.json \\
    --endpoints single-api/endpoints-Amazon.Advertising.Api.json \\
    --swagger-title "Amazon.Advertising.Api" \\
    --module  "SupplementData" \\
    --env     us

写入两个目标：
  1. [模块汇总] sheet：Case数 / 通过率 / 接口覆盖率
  2. [Amazon.Advertising.Api] / [PacvueMainApi] 等 swagger sheet：场景覆盖列
"""
import argparse, json, re, os, glob
from collections import defaultdict
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

EXCEL_PATH = None  # 从 cases 路径自动推断：single-api/<服务>/swagger_modules.xlsx

# ── helpers ──────────────────────────────────────────────────────────────────
def _thin_border():
    s = Side(style='thin', color='BFBFBF')
    return Border(left=s, right=s, top=s, bottom=s)

def _find_or_create_col(ws, col_name, header_font=None, header_fill=None, width=None):
    for c in range(1, ws.max_column + 1):
        if ws.cell(1, c).value == col_name:
            return c
    col_idx = ws.max_column + 1
    cell = ws.cell(1, col_idx, col_name)
    if header_font:  cell.font  = header_font
    if header_fill:  cell.fill  = header_fill
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border    = _thin_border()
    if width:
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    return col_idx

def norm_path(p, base_url=''):
    """把 cases.json 里的相对 path（不含/api前缀，走ROOTURL的接口也不含）补全成
    与 endpoints-*.json 的 path 同格式的绝对路径，便于跨数据源匹配。"""
    if not p:
        return p
    p = p.split('?', 1)[0]
    if p.startswith('/api/') or p.startswith('/'):
        return p if p.startswith('/api/') or 'ROOTURL' in base_url else '/api' + p
    if 'ROOTURL' in base_url:
        return '/' + p
    return '/api/' + p

def _canon(p):
    # 归一化：去 query、去 /api/ 前缀、数字ID/{占位} 段统一为 * ，以便按端点模板匹配
    if not p:
        return ''
    p = p.split('?', 1)[0]
    p = p.split('/api/', 1)[-1] if '/api/' in p else p.lstrip('/')
    segs = ['*' if (s.isdigit() or (s.startswith('{') and s.endswith('}'))) else s.lower()
            for s in p.split('/')]
    return '/'.join(segs)

# ── Phase 5.1 — 模块汇总 ─────────────────────────────────────────────────────
def update_summary(wb, swagger_title, module, env, cases, report, endpoints):
    ws = wb['模块汇总']

    module_eps  = [e for e in endpoints if e.get('tag') == module]
    api_total   = len(module_eps)
    # cases.json 的 step 只有相对 path（不含/api前缀），先用 norm_path 按 base_url 补全成
    # 绝对路径，再归一化按 (方法, 端点模板) 去重匹配；report.json 里的 step 只有完整 url
    # （含域名/服务前缀），不能直接拿来做路径匹配，因此改用 cases.json 作为匹配数据源。
    case_keys   = {((s.get('method') or '').upper(), _canon(norm_path(s.get('path', ''), s.get('base_url', ''))))
                   for c in cases for s in c.get('steps', [])}
    api_covered = len({((e.get('method') or '').upper(), _canon(e.get('path', ''))) for e in module_eps
                       if ((e.get('method') or '').upper(), _canon(e.get('path', ''))) in case_keys})

    summary_font = Font(bold=True)
    summary_fill = PatternFill('solid', fgColor='D9E1F2')

    # 找各列位置（按列名，顺序不固定）
    col_idx = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(1, c).value or ''
        if '环境' in v:        col_idx['env']      = c
        if 'Swagger' in v:    col_idx['swagger']   = c
        if 'Platform' in v:   col_idx['platform']  = c  # rule-api 多平台结构
        if 'Tag' in v:        col_idx['module']    = c
    for name in ('Case数', '通过率', '接口覆盖率'):
        col_idx[name] = _find_or_create_col(ws, name, summary_font, summary_fill)

    # 兼容两种结构：标准服务用 swagger 列，多平台服务（rule-api）用 platform 列
    has_swagger  = 'swagger'  in col_idx
    has_platform = 'platform' in col_idx

    total  = report['total']
    passed = report['passed']
    pass_rate = (str(round(passed / total * 100)) + '%') if total else 'N/A'
    coverage  = (str(round(api_covered / api_total * 100)) + '%'
                 + ' (' + str(api_covered) + '/' + str(api_total) + ')') if api_total else 'N/A'

    for r in range(2, ws.max_row + 1):
        # 环境列常为合并单元格（仅块首行有值，其余为 None）；None 视为继承块环境，不作否决
        env_val = ws.cell(r, col_idx['env']).value if 'env' in col_idx else None
        env_ok = env_val in (env, None, '')
        mod_ok  = 'module' in col_idx and ws.cell(r, col_idx['module']).value == module
        if has_swagger:
            key_ok = ws.cell(r, col_idx['swagger']).value == swagger_title
        elif has_platform:
            # 多平台服务：swagger_title 传平台名（如 tiktok）
            key_ok = ws.cell(r, col_idx['platform']).value == swagger_title
        else:
            key_ok = True  # 无 swagger/platform 列时仅按 module 匹配
        if env_ok and key_ok and mod_ok:
            ws.cell(r, col_idx['Case数'],    total).alignment    = Alignment(horizontal='center')
            ws.cell(r, col_idx['通过率'],    pass_rate).alignment = Alignment(horizontal='center')
            ws.cell(r, col_idx['接口覆盖率'], coverage).alignment  = Alignment(horizontal='center')
            print(f'[模块汇总] {swagger_title}/{module}[{env}]: case={total}, pass={pass_rate}, coverage={coverage}')
            return
    print(f'[模块汇总] WARNING: row not found for {swagger_title}/{module}[{env}]')

# ── Phase 5.2 — 接口级场景覆盖（单列「场景覆盖」，多行文本）────────────────────
def update_scenario_col(wb, swagger_sheet, env, cases, report):
    if swagger_sheet not in wb.sheetnames:
        print(f'[场景覆盖] sheet "{swagger_sheet}" not found, skip')
        return

    ws  = wb[swagger_sheet]
    bdr = _thin_border()
    h_font = Font(bold=True, color='FFFFFF')
    h_fill = PatternFill('solid', fgColor='2E5F8A')

    norm = norm_path

    # Build (method, path) → scenario lines（从 cases 的 name/description 提取）
    method_path_scenarios = defaultdict(list)
    for c in cases:
        steps = c.get('steps', [])
        if not steps:
            continue
        # 优先从 case name 提取方法和路径（格式："{METHOD} /api/xxx - desc"）
        # 避免多步骤 case 的前置步骤路径干扰归属
        # 【高风险-仅生成不执行】等前缀会挡在 METHOD 前面，先去掉再匹配，
        # 否则会 fallback 到 steps[-1]（多为后置清理步骤），把场景挂到错误的接口上
        name_no_prefix = re.sub(r'^【[^】]*】', '', c.get('name', ''))
        name_m = re.match(r'^([A-Z]+)\s+(/\S+)', name_no_prefix)
        if name_m:
            case_method = name_m.group(1).upper()
            raw_path    = name_m.group(2)
            ref_step    = steps[0]
        else:
            # fallback: 取最后一步（实际被测接口）
            last = steps[-1]
            case_method = (last.get('method') or '').upper()
            raw_path    = last.get('path', '')
            ref_step    = last
        full_path = norm(raw_path, ref_step.get('base_url', ''))
        name = c.get('name', '')
        desc = c.get('description', '')
        label_raw = name.split(' - ', 1)[1] if ' - ' in name else name
        m = re.match(r'^(场景\d+)[_\-](.+)$', label_raw)
        label = (m.group(1) + ': ' + m.group(2).replace('_', ' ')) if m else label_raw.replace('_', ' ')
        pct_m = re.search(r'占(?:比)?(?:约)?\s*([\d.]+%)', desc)
        if pct_m:
            entry = label + ' — ' + pct_m.group(1)
        elif '无ES流量' in desc or 'xx%' in desc:
            entry = label + '（无ES流量）'
        else:
            entry = label
        method_path_scenarios[(case_method, full_path)].append(entry)

    # 找列位置：方法列 + 接口路径列 + 既有「场景覆盖」列
    env_col  = next((c for c in range(1, ws.max_column + 1) if '环境' in str(ws.cell(1, c).value or '')), None)
    path_col = next((c for c in range(1, ws.max_column + 1)
                     if '路径' in str(ws.cell(1, c).value or '') or
                        'path' in str(ws.cell(1, c).value or '').lower()), None)
    method_col = next((c for c in range(1, ws.max_column + 1)
                       if str(ws.cell(1, c).value or '').strip().lower() in ('method', '方法', 'http方法')), None)
    if path_col is None:
        print(f'[场景覆盖] path column not found in {swagger_sheet}'); return

    scene_col = _find_or_create_col(ws, '场景覆盖', h_font, h_fill, width=72)

    # 归一化索引：(method, canon_path) → entries
    canon_scenarios = {(meth, _canon(p)): v for (meth, p), v in method_path_scenarios.items()}

    updated = 0
    for r in range(2, ws.max_row + 1):
        if env_col and ws.cell(r, env_col).value not in (env, None, ''):
            continue
        path = ws.cell(r, path_col).value
        if not path:
            continue
        row_method = (ws.cell(r, method_col).value or '').strip().upper() if method_col else ''
        # 优先按 (method, path) 精确匹配；无方法列时退化为只按路径匹配
        lines = None
        if row_method:
            lines = (method_path_scenarios.get((row_method, path))
                     or canon_scenarios.get((row_method, _canon(path))))
        if lines is None:
            # 兼容无方法列的 sheet：合并所有方法的场景
            all_lines = []
            for (meth, p), v in method_path_scenarios.items():
                if p == path or _canon(p) == _canon(path):
                    all_lines.extend(v)
            lines = all_lines or None
        if lines:
            cell = ws.cell(r, scene_col, '\n'.join(lines))
            cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
            cell.border    = bdr
            ws.row_dimensions[r].height = max(ws.row_dimensions[r].height or 15, len(lines) * 16)
            updated += 1

    print(f'[场景覆盖] {swagger_sheet}[{env}]: updated {updated} rows')

# ── 跨轮次累计 ─────────────────────────────────────────────────────────────────
def _module_dir_from_cases_path(cases_path):
    # single-api/mainapi/<env>/<swagger>/<module>/task-*/cases.json -> .../<module>
    return os.path.dirname(os.path.dirname(os.path.abspath(cases_path)))

def load_all_tasks_for_module(cases_path, fallback_cases, fallback_report):
    """一个模块通常会分多轮（多个 task-* 目录）生成 case。模块汇总的 Case数/通过率/
    接口覆盖率如果只用本次传入的单轮 cases/report 计算，会被最后一轮覆盖，看不出
    真实累计进度。这里改为扫描该模块目录下所有 task-*/cases.json + report.json，
    合并后再统计，使结果与实际累计覆盖一致。"""
    module_dir = _module_dir_from_cases_path(cases_path)
    task_dirs = sorted(glob.glob(os.path.join(module_dir, 'task-*')))
    all_cases, total, passed, failed, skipped = [], 0, 0, 0, 0
    found_any = False
    for td in task_dirs:
        cp = os.path.join(td, 'cases.json')
        rp = os.path.join(td, 'report.json')
        if not (os.path.exists(cp) and os.path.exists(rp)):
            continue
        with open(cp, encoding='utf-8-sig') as f: c = json.load(f)
        with open(rp, encoding='utf-8-sig') as f: r = json.load(f)
        all_cases.extend(c)
        total   += r.get('total', len(c))
        passed  += r.get('passed', 0)
        failed  += r.get('failed', 0)
        skipped += r.get('skipped', 0)
        found_any = True
    if not found_any:
        return fallback_cases, fallback_report
    merged_report = {'total': total, 'passed': passed, 'failed': failed, 'skipped': skipped}
    return all_cases, merged_report

# ── main ─────────────────────────────────────────────────────────────────────
def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--cases',          required=True)
    ap.add_argument('--report',         required=True)
    ap.add_argument('--endpoints',      required=True)
    ap.add_argument('--swagger-title',  required=True, dest='swagger_title')
    ap.add_argument('--sheet-name',     default=None,  dest='sheet_name',
                    help='场景覆盖 sheet 名，默认与 swagger-title 相同；多平台服务可单独指定（如 RuleApi）')
    ap.add_argument('--module',         required=True)
    ap.add_argument('--env',            default='us', choices=['us', 'cn', 'eu'])
    ap.add_argument('--excel',          default=None,
                    help='Excel 路径，默认从 cases 路径推断：single-api/<服务>/swagger_modules.xlsx')
    args = ap.parse_args()
    if not args.sheet_name:
        args.sheet_name = args.swagger_title

    # 自动推断 excel 路径：取 cases 路径中 single-api/<服务> 部分
    if not args.excel:
        import re as _re
        m = _re.match(r'(.*?single-api/[^/]+)/', args.cases.replace('\\', '/'))
        args.excel = (m.group(1) + '/swagger_modules.xlsx') if m else 'single-api/swagger_modules.xlsx'

    with open(args.cases,     encoding='utf-8-sig') as f: cases     = json.load(f)
    with open(args.report,    encoding='utf-8-sig') as f: report    = json.load(f)
    with open(args.endpoints, encoding='utf-8-sig') as f: endpoints = json.load(f)

    cases, report = load_all_tasks_for_module(args.cases, cases, report)

    wb = openpyxl.load_workbook(args.excel)
    update_summary(wb, args.swagger_title, args.module, args.env, cases, report, endpoints)
    update_scenario_col(wb, args.sheet_name, args.env, cases, report)
    # 原文件常被 Excel 打开占用：先存 tmp，再尝试替换；替换失败则保留 tmp 供手动合并
    import os, shutil
    tmp = args.excel.replace('.xlsx', '_tmp.xlsx')
    wb.save(tmp)
    try:
        shutil.move(tmp, args.excel)
        print('Excel saved:', args.excel)
    except PermissionError:
        print('Excel 被占用（请关闭后手动用 tmp 覆盖）。tmp 已保存:', tmp)

if __name__ == '__main__':
    main()
