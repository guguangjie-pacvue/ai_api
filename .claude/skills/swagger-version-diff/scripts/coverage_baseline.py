"""
coverage_baseline.py — 版本变更接口测试覆盖基准（对齐 swagger）

以 case-design-<commit>.json 为设计基准，合并 report.json 的实际执行结果，
渲染成一行一接口的 Excel 覆盖表，直观对应 swagger 影响范围。

Usage:
  python coverage_baseline.py \
    --design single-api/rule-api/kevel/us/case-design-1f18987.json \
    --report single-api/rule-api/kevel/us/smoke/task-xxx/report-1f18987.json \
    [--report ...更多 report，可多次] \
    --out    single-api/rule-api/kevel/coverage-1f18987.xlsx

每行一个接口，列：
  Method | Path | 模块 | 变更类型 | 设计case数 | 已执行 | 通过 | 失败 | 阻塞 | 覆盖率 | 状态 | 卡点/失败原因
状态色：全通过=绿，有失败=红，全阻塞=灰，部分=黄
"""
import argparse, json, re
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HEADERS = ['Method', 'Path', '模块', '变更类型', '设计case数', '已执行', '通过', '失败', '阻塞', '覆盖率', '状态', '卡点/失败原因']

def _border():
    s = Side(style='thin', color='BFBFBF')
    return Border(left=s, right=s, top=s, bottom=s)

def _canon(p):
    if not p:
        return ''
    p = p.split('?', 1)[0].lstrip('/')
    segs = ['*' if (s.isdigit() or (s.startswith('{') and s.endswith('}'))) else s.lower()
            for s in p.split('/')]
    return '/'.join(segs)

def load_report_status(report_paths):
    """(method, canon_path) -> {'pass': n, 'fail': n} 汇总所有 report 的 step 结果。"""
    agg = {}
    for rp in report_paths:
        with open(rp, encoding='utf-8-sig') as f:
            rep = json.load(f)
        for case in rep.get('cases', []):
            for st in case.get('steps', []):
                m = (st.get('method') or '').upper()
                path = st.get('path') or st.get('url') or ''
                # report 的 url 是完整地址，取 path 段
                path = re.sub(r'^https?://[^/]+', '', path)
                key = (m, _canon(path))
                d = agg.setdefault(key, {'pass': 0, 'fail': 0})
                if st.get('passed') or st.get('status') == 'PASS':
                    d['pass'] += 1
                else:
                    d['fail'] += 1
    return agg

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--design', required=True)
    ap.add_argument('--report', action='append', default=[])
    ap.add_argument('--out', required=True)
    args = ap.parse_args()

    with open(args.design, encoding='utf-8-sig') as f:
        design = json.load(f)
    rep_status = load_report_status(args.report)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'coverage-' + str(design.get('version', ''))

    h_font = Font(bold=True, color='FFFFFF')
    h_fill = PatternFill('solid', fgColor='2E5F8A')
    for c, name in enumerate(HEADERS, 1):
        cell = ws.cell(1, c, name)
        cell.font = h_font; cell.fill = h_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = _border()

    widths = [8, 42, 12, 12, 10, 8, 6, 6, 6, 10, 10, 48]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    green = PatternFill('solid', fgColor='C6EFCE')
    red   = PatternFill('solid', fgColor='FFC7CE')
    grey  = PatternFill('solid', fgColor='D9D9D9')
    yellow= PatternFill('solid', fgColor='FFEB9C')

    r = 2
    tot_design = tot_run = tot_pass = tot_fail = tot_block = 0
    for ep in design['endpoints']:
        cases = ep.get('cases', [])
        design_n = len(cases)
        # 设计态统计
        blocked = [c for c in cases if c.get('status') == 'BLOCKED']
        # 实际执行结果（优先用 report 覆盖设计态）
        key = ((ep.get('method') or '').upper(), _canon(ep.get('path', '')))
        actual = rep_status.get(key)
        if actual:
            run_n = actual['pass'] + actual['fail']
            pass_n = actual['pass']; fail_n = actual['fail']
        else:
            run_n = len([c for c in cases if c.get('status') in ('PASS', 'FAIL')])
            pass_n = len([c for c in cases if c.get('status') == 'PASS'])
            fail_n = len([c for c in cases if c.get('status') == 'FAIL'])
        block_n = len(blocked)
        cov = (str(round(run_n / design_n * 100)) + '%') if design_n else 'N/A'

        if design_n and pass_n == design_n:
            status, fill = '全通过', green
        elif fail_n > 0:
            status, fill = '有失败', red
        elif block_n == design_n and design_n:
            status, fill = '全阻塞', grey
        elif run_n > 0:
            status, fill = '部分', yellow
        else:
            status, fill = '未测', grey

        blockers = sorted({c.get('blocker', '') for c in blocked if c.get('blocker')})
        reason = '; '.join(blockers)

        row = [ep.get('method'), ep.get('path'), ep.get('module'), ep.get('change_type'),
               design_n, run_n, pass_n, fail_n, block_n, cov, status, reason]
        for c, v in enumerate(row, 1):
            cell = ws.cell(r, c, v)
            cell.border = _border()
            cell.alignment = Alignment(horizontal='left' if c in (2, 12) else 'center',
                                       vertical='center', wrap_text=(c == 12))
            if c == 11:
                cell.fill = fill
        r += 1
        tot_design += design_n; tot_run += run_n; tot_pass += pass_n
        tot_fail += fail_n; tot_block += block_n

    # 合计行
    tot = ['合计', f'{len(design["endpoints"])} 接口', '', '', tot_design, tot_run, tot_pass,
           tot_fail, tot_block, (str(round(tot_run/tot_design*100))+'%') if tot_design else 'N/A', '', '']
    for c, v in enumerate(tot, 1):
        cell = ws.cell(r, c, v)
        cell.font = Font(bold=True); cell.fill = PatternFill('solid', fgColor='D9E1F2')
        cell.border = _border()
        cell.alignment = Alignment(horizontal='left' if c == 2 else 'center', vertical='center')

    ws.freeze_panes = 'A2'
    wb.save(args.out)
    print(f'覆盖基准已生成: {args.out}')
    print(f'  接口 {len(design["endpoints"])} | 设计 case {tot_design} | 已执行 {tot_run} | 通过 {tot_pass} | 失败 {tot_fail} | 阻塞 {tot_block}')

if __name__ == '__main__':
    main()
